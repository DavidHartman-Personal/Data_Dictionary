"""Create a data dictionary object from an Excel Workbook that contains data/details.

This script data from an Excel workbook to create a data dictionary object.

TODO: Add function to confirm which row contains the headers for the Entity/Attributes.

"""

import argparse
import configparser
import os
import logging
import coloredlogs
from data_dictionary import constants
from openpyxl import load_workbook, worksheet
import pandas as pd
import json
import pprint as pp
from model import DataDictionary as dd
from pathlib import Path
import numpy as np

coloredlogs.install(level=logging.DEBUG,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')


def create_entities_from_excel_worksheet(spreadsheet_file: Path, data_dictionary: dd):
    """Creates Entity object from an Excel Worksheet and adds it to the data dictionary

    This function takes in a worksheet that is in the appropriate format and creates an Entity object
    that is then added to the data dictionary.  This function uses panda data frames to read the Excel file.
    TODO: Add check to make sure worksheet exists in Excel file.
    TODO: Add check that worksheet contains all required header fields (e.g. NAME, DESCRIPTION, etc.)

    Args:
      data_dictionary (DataDictionary):
      spreadsheet_file (Path): The worksheet name

    Returns:
      None
    """
    logging.info("Creating entities from excel worksheet")
    excel_poib_data = pd.read_excel(spreadsheet_file,
                                    sheet_name=constants.SOURCE_DATA_DICT_MAPPING['source_worksheet_name']).replace(
        np.nan, None)
    entity = dd.Entity(name=constants.SOURCE_DATA_DICT_MAPPING['entity_name'],
                       description=constants.SOURCE_DATA_DICT_MAPPING['entity_name'],
                       subject_area=constants.SOURCE_DATA_DICT_MAPPING['subject_area'],
                       environment=constants.SOURCE_DATA_DICT_MAPPING['environment']
                       )
    for row_label, row in excel_poib_data.iterrows():
        entity_attribute = dd.Attribute(
            name=constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['name'],
            description=constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['description'],
            data_type=constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['data_type'],
            subject_area=constants.SOURCE_DATA_DICT_MAPPING['subject_area'],
            environment=constants.SOURCE_DATA_DICT_MAPPING['environment'],
            max_length=constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['max_length']
        )
        entity.add_attribute(entity_attribute)
    data_dictionary.add_entity(entity)


def read_in_data_dictionary(data_dictionary_source_file: Path) -> dd:
    """Creates DataDictionary instance from a JSON file

    Args:
        data_dictionary_source_file (Path): Location of existing data dictionary JSON File.
    Returns:
         DataDictionary: Data dictionary object
    """
    logging.info("Reading in data dictionary...")
    data_dict_from_json = dd.DataDictionary.create_data_dictionary_from_json(data_dictionary_source_file)
    return data_dict_from_json


def create_data_dictionary_from_json_file():
    input_file_name: str = "po_sample.xlsx"
    output_file_name: str = "data_dictionary.json"
    dd_output_file: Path = constants.OUTPUT_FOLDER / output_file_name

    dd_1 = read_in_data_dictionary(constants.TEST_DATA_FILE)
    dd_1.write_data_dictionary(dd_output_file)
    dd_1.create_markdown_files(constants.MARKDOWN_FILE_DIR, force_overwrite=True)


def create_data_dictionary_from_config_data():
    """

    """
    # create a base DataDictionary instance
    data_dict = dd.DataDictionary(name="USAID_Data_Dictionary",
                                  description="USAID Data Dictionary for One Network system",
                                  subject_area="All",
                                  environment="All"
                                  )
    for entity in constants.ENTITIES_SOURCES:
        logging.info("entities source definition: [%s] excel file/worksheet: [%s] - [%s]",
                     str(entity['entity_name']),
                     entity['source_file_name'],
                     entity['source_worksheet_name'])
        new_entity = dd.Entity(name=entity['entity_name'],
                           description=entity['description'],
                           subject_area=entity['subject_area'],
                           environment=entity['environment']
                           )
        excel_poib_data = pd.read_excel(entity['source_file_name'],
                                        sheet_name=entity['source_worksheet_name']).replace(np.nan, None)
        for row_label, row in excel_poib_data.iterrows():
            logging.info("Row label: [%s] Row values: [%s]", str(row_label), str(row))
            # entity_attribute = dd.Attribute(
            #     name=entity['attribute_column_mapping']['name'],
            #     description=entity['attribute_column_mapping']['description'],
            #     data_type=entity['attribute_column_mapping']['data_type'],
            #     subject_area=entity['subject_area'],
            #     environment=entity['environment'],
            #     max_length=entity['attribute_column_mapping']['max_length']
            # )
            # entity.add_attribute(entity_attribute)
        # print entity defined and excel worksheet source for entity/attribute definitions
        # create_entities_from_excel_worksheet(entity['source_file_name'], data_dict)


def main():
    """The main method for this script.

    """

    def process_args():
        logging.info("Processing command line Arguments")
        parser = argparse.ArgumentParser(description=__doc__)
        parser.add_argument('input_file', type=str, help="The spreadsheet file to print the columns of",
                            nargs='?', default=constants.DEFAULT_SPREADSHEET)
        parser.add_argument('--config', '-c', dest='filename_config', action='store',
                            default=constants.FILENAME_INPUT_CONFIG,
                            help='Config file for generating AWS User Access Report - Default is Environment variable CONFIG_FILE_PATH')
        args = parser.parse_args()
        return args.filename_config

    config_filename = process_args()

    def get_configs(config_filename_in):
        logging.info("Getting config file values from [%s]", str(config_filename_in))
        config = configparser.ConfigParser()
        try:
            with open(config_filename_in) as f:
                config.read(config_filename)
        except Exception as e:
            logging.error("Error Reading config file [%s]: [%s]", config_filename, str(e))
            raise ValueError("Error Reading config file [%s]: [%s]", config_filename, str(e))
        header_row_out = config.get('po_sample', 'header_row', fallback=constants.DEFAULT_HEADER_ROW)
        first_col_out = config.get('po_sample', 'first_col', fallback=constants.DEFAULT_FIRST_COL)
        last_col_out = config.get('po_sample', 'last_col', fallback=constants.DEFAULT_LAST_COL)
        source_file_name_out = config.get('po_sample', 'source_file_name', fallback=constants.DEFAULT_SPREADSHEET)
        source_worksheet_name_out = config.get('po_sample', 'source_worksheet_name')
        return header_row_out, first_col_out, last_col_out, source_file_name_out, source_worksheet_name_out

    (header_row, first_col, last_col, source_spreadsheet_file, source_worksheet_name) = get_configs(config_filename)
    input_file_name = "po_sample.xlsx"
    output_file_name = "data_dictionary.json"
    test_spreadsheet = constants.SOURCE_FOLDER / input_file_name
    dd_output_file = constants.OUTPUT_FOLDER / output_file_name
    data_dict = dd.DataDictionary(name="USAID_Data_Dictionary",
                                  description="USAID Data Dictionary for One Network system",
                                  subject_area="All",
                                  environment="All",
                                  source_filename=test_spreadsheet
                                  )
    create_entities_from_excel_worksheet(test_spreadsheet, data_dict)
    data_dict.write_data_dictionary(dd_output_file)


if __name__ == "__main__":
    # main()
    # create_data_dictionary_from_json_file()
    create_data_dictionary_from_config_data()

    # excel_poib_data = pd.read_excel(test_spreadsheet,
    #                                 sheet_name=constants.SOURCE_DATA_DICT_MAPPING['source_worksheet_name']).replace(
    #     np.nan, None)
    # entity = dd.Entity(name=constants.SOURCE_DATA_DICT_MAPPING['entity_name'],
    #                    description=constants.SOURCE_DATA_DICT_MAPPING['entity_name'],
    #                    subject_area=constants.SOURCE_DATA_DICT_MAPPING['subject_area'],
    #                    environment=constants.SOURCE_DATA_DICT_MAPPING['environment']
    #                    )
    # for row_label, row in excel_poib_data.iterrows():
    #     entity_attribute = dd.Attribute(
    #         name=row[constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['name']],
    #         description=row[constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['description']],
    #         data_type=row[constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['data_type']],
    #         subject_area=constants.SOURCE_DATA_DICT_MAPPING['subject_area'],
    #         environment=constants.SOURCE_DATA_DICT_MAPPING['environment'],
    #         max_length=row[constants.SOURCE_DATA_DICT_MAPPING['attribute_column_mapping']['max_length']],
    #     )
    #     entity.add_attribute(entity_attribute)
    # data_dict.add_entity(entity)
    # def get_all_tables(filename):
    #     """ Get all tables from a given workbook. Returns a dictionary of tables.
    #         Requires a filename, which includes the file path and filename.
    #
    #     Args:
    #         filename (str): The Excel filename
    #
    #     Returns:
    #         object:
    #     """
    #
    #     # Load the workbook, from the filename, setting read_only to False
    #     wb = load_workbook(filename=filename, read_only=False, keep_vba=False, data_only=True, keep_links=False)
    #
    #     # Initialize the dictionary of tables
    #     tables_dict = {}
    #
    #     # Go through each worksheet in the workbook
    #     for ws_name in wb.sheetnames:
    #         print("")
    #         print(f"worksheet name: {ws_name}")
    #         ws = wb[ws_name]
    #         print(f"tables in worksheet: {len(ws.tables)}")
    #
    #         # Get each table in the worksheet
    #         for tbl in ws.tables.values():
    #             print(f"table name: {tbl.name}")
    #             # First, add some info about the table to the dictionary
    #             tables_dict[tbl.name] = {
    #                 'table_name': tbl.name,
    #                 'worksheet': ws_name,
    #                 'num_cols': len(tbl.tableColumns),
    #                 'table_range': tbl.ref}
    #
    #             # Grab the 'data' from the table
    #             data = ws[tbl.ref]
    #
    #             # Now convert the table 'data' to a Pandas DataFrame
    #             # First get a list of all rows, including the first header row
    #             rows_list = []
    #             for row in data:
    #                 print(str(row))
    #                 # Get a list of all columns in each row
    #                 cols = []
    #                 for col in row:
    #                     cols.append(col.value)
    #                 rows_list.append(cols)
    #
    #             # Create a pandas dataframe from the rows_list.
    #             # The first row is the column names
    #             df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
    #             print(str(df))
    #
    #             # Add the dataframe to the dictionary of tables
    #             tables_dict[tbl.name]['dataframe'] = df
    #
    #     return tables_dict
